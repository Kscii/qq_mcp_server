from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

import qq_mcp_server.cards as cards_module
from qq_mcp_server.cards import FixedCharacterCardParser


def make_card(path: Path, *, name: str = "调查员", player: str = "玩家") -> Path:
    book = Workbook()
    sheet = book.active
    sheet.title = "人物卡"
    sheet["J3"] = "玩家"
    sheet["F16"] = "技能名称"
    sheet["F79"] = "物品名称"
    sheet["E3"] = name
    sheet["M3"] = player
    sheet["E4"] = 30
    sheet["M4"] = "现代"
    sheet["E5"] = "记者"
    sheet["E6"] = "女"
    sheet["M6"] = "中国"
    sheet["E7"] = "上海"
    sheet["M7"] = "杭州"
    for coordinate, value in {
        "U3": 50,
        "AA3": 60,
        "AG3": 55,
        "U5": 50,
        "AA5": 60,
        "AG5": 70,
        "U7": 50,
        "AA7": 65,
        "AG7": 55,
    }.items():
        sheet[coordinate] = value
    sheet["E10"] = 10
    sheet["G10"] = 10
    sheet["N10"] = 55
    sheet["P10"] = 99
    sheet["W10"] = 11
    sheet["Y10"] = 11
    sheet["F17"] = "侦查"
    sheet["J17"] = 25
    sheet["N17"] = 45
    sheet["R17"] = 70
    sheet["T17"] = 35
    sheet["V17"] = 14
    sheet["B54"] = "手枪"
    sheet["G54"] = "射击"
    sheet["M54"] = "手枪"
    sheet["Q54"] = 50
    sheet["W54"] = "1D10"
    sheet["AA62"] = "总戴着旧呢帽。"
    sheet["AA66"] = "大学导师"
    sheet["F80"] = "录音笔"
    sheet["D80"] = "衣袋"
    sheet["B80"] = "隐藏"
    other = book.create_sheet("不应读取")
    other["A1"] = "这张表的内容不能进入人物卡"
    book.save(path)
    return path


def test_fixed_card_parses_all_meaningful_sections_and_ignores_other_sheets(
    tmp_path: Path,
) -> None:
    card = FixedCharacterCardParser().parse(make_card(tmp_path / "card.xlsx")).document
    assert card["identity"]["name"] == "调查员"
    assert card["vitals"]["hp"] == {"current": 10, "max": 10}
    assert card["skills"]["侦查"]["regular"] == 70
    assert card["weapons"][0]["name"] == "手枪"
    assert card["background"]["important_people"] == "大学导师"
    assert card["inventory"][0]["name"] == "录音笔"
    assert "不应读取" not in str(card)


def test_fixed_card_loads_one_workbook_and_streams_cached_values_once(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = make_card(tmp_path / "card.xlsx")
    real_load_workbook = cards_module.load_workbook
    calls: list[dict[str, object]] = []

    def tracked_load_workbook(*args: object, **kwargs: object) -> object:
        calls.append(kwargs)
        return real_load_workbook(*args, **kwargs)

    monkeypatch.setattr(cards_module, "load_workbook", tracked_load_workbook)
    card = FixedCharacterCardParser().parse(path)

    assert card.character_name == "调查员"
    assert calls == [{"read_only": True, "data_only": True, "keep_links": False}]


def test_blank_template_is_recognized_but_cannot_be_bound(tmp_path: Path) -> None:
    path = make_card(tmp_path / "blank.xlsx", name="", player="")
    with pytest.raises(ValueError, match="姓名和玩家不能为空"):
        FixedCharacterCardParser().parse(path)


def test_real_blank_blue_template_when_available() -> None:
    path = Path("/home/kscii/Downloads/COC空白卡浅蓝色.xlsx")
    if not path.is_file():
        pytest.skip("本机真实空白人物卡样本不存在")
    with pytest.raises(ValueError, match="姓名和玩家不能为空"):
        FixedCharacterCardParser().parse(path)


def test_unknown_template_is_rejected(tmp_path: Path) -> None:
    book = Workbook()
    book.active.title = "人物卡"
    path = tmp_path / "unknown.xlsx"
    book.save(path)
    with pytest.raises(ValueError, match="不是已锁定"):
        FixedCharacterCardParser().parse(path)


@pytest.mark.parametrize(
    ("filename", "name", "important"),
    [
        ("山田何也(1).xlsx", "山田何也", "山田奈绪"),
        ("小林和彦(2).xlsx", "小林和彦", None),
    ],
)
def test_real_fixed_cards_when_available(filename: str, name: str, important: str | None) -> None:
    path = Path("/home/kscii/Downloads") / filename
    if not path.is_file():
        pytest.skip("本机真实人物卡样本不存在")
    card = FixedCharacterCardParser().parse(path).document
    assert card["identity"]["name"] == name
    assert len(card["skills"]) >= 60
    if important:
        assert important in card["background"]["important_people"]
