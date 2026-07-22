from __future__ import annotations

import copy
import hashlib
import os
import re
import secrets
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Any, cast

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from qq_mcp_server.store import MessageStore

TEMPLATE_ID = "beier_black_gold_23_1_1"


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _text(value: object) -> str | None:
    if value is None:
        return None
    result = str(value).strip()
    if not result or result.startswith("#"):
        return None
    return result


def _integer(value: object) -> int | None:
    if value is None or value == "" or isinstance(value, bool):
        return None
    try:
        return int(float(str(value)))
    except ValueError:
        return None


def _slug(value: str) -> str:
    normalized = re.sub(r"[^\w]+", "_", value.strip().lower(), flags=re.UNICODE).strip("_")
    return normalized or "character"


def _skill_name(base_name: str | None, custom_name: str | None) -> str | None:
    if not base_name:
        return custom_name
    if not custom_name:
        return base_name
    category = base_name.rstrip("：:")
    if base_name.endswith(("：", ":")) or re.search(r"[①②③]$", base_name):
        return f"{category}（{custom_name}）"
    return custom_name


def _prune(value: Any) -> Any:
    if isinstance(value, dict):
        return {
            key: cleaned
            for key, item in value.items()
            if (cleaned := _prune(item)) not in (None, "", [], {})
        }
    if isinstance(value, list):
        return [cleaned for item in value if (cleaned := _prune(item)) not in (None, "", [], {})]
    return value


@dataclass(frozen=True, slots=True)
class ParsedCard:
    document: dict[str, Any]
    source_sha256: str
    source_filename: str

    @property
    def character_name(self) -> str:
        return str(self.document["identity"]["name"])


class FixedCharacterCardParser:
    """只识别已经锁定的浅蓝色 23.1.1 人物卡。"""

    @staticmethod
    def _verify(sheet: Worksheet) -> None:
        anchors = {"J3": "玩家", "F16": "技能名称", "F79": "物品名称"}
        mismatches = [
            f"{coordinate}={sheet[coordinate].value!r}"
            for coordinate, expected in anchors.items()
            if _text(sheet[coordinate].value) != expected
        ]
        if mismatches:
            raise ValueError("人物卡不是已锁定的浅蓝色 23.1.1 模板：" + "，".join(mismatches))

    def parse(self, source: Path) -> ParsedCard:
        source = source.expanduser().resolve()
        if not source.is_file() or source.suffix.lower() != ".xlsx":
            raise ValueError("人物卡必须是可读取的 .xlsx 文件")
        if source.stat().st_size > 16 * 1024 * 1024:
            raise ValueError("人物卡文件不能超过 16 MiB")

        formula_book = load_workbook(source, read_only=True, data_only=False)
        value_book = load_workbook(source, read_only=True, data_only=True)
        try:
            if "人物卡" not in formula_book.sheetnames or "人物卡" not in value_book.sheetnames:
                raise ValueError("工作簿缺少 人物卡 工作表")
            formula_sheet = formula_book["人物卡"]
            sheet = value_book["人物卡"]
            self._verify(formula_sheet)
            document = self._parse_sheet(sheet)
        finally:
            formula_book.close()
            value_book.close()
        return ParsedCard(document, _sha256(source), source.name)

    def _parse_sheet(self, sheet: Worksheet) -> dict[str, Any]:
        provenance: dict[str, list[str]] = {}

        def value(coordinate: str, pointer: str) -> object:
            provenance.setdefault(pointer, []).append(coordinate)
            return sheet[coordinate].value

        identity_cells = {
            "name": "E3",
            "player": "M3",
            "occupation": "E5",
            "age": "E4",
            "era": "M4",
            "sex": "E6",
            "nationality": "M6",
            "residence": "E7",
            "birthplace": "M7",
        }
        identity: dict[str, Any] = {}
        for key, coordinate in identity_cells.items():
            raw = value(coordinate, f"/identity/{key}")
            identity[key] = _integer(raw) if key == "age" else _text(raw)
        if not identity["name"] or not identity["player"]:
            raise ValueError("人物卡姓名和玩家不能为空")

        attribute_cells = {
            "str": "U3",
            "dex": "AA3",
            "pow": "AG3",
            "con": "U5",
            "app": "AA5",
            "edu": "AG5",
            "siz": "U7",
            "int": "AA7",
            "luck": "AG7",
        }
        attributes: dict[str, int] = {}
        for key, coordinate in attribute_cells.items():
            parsed = _integer(value(coordinate, f"/attributes/{key}"))
            if parsed is None:
                raise ValueError(f"必填属性 {key.upper()} 缺失（{coordinate}）")
            attributes[key] = parsed

        era_time = {
            "calendar": _text(value("E8", "/era_time/calendar")),
            "year": _integer(value("G8", "/era_time/year")),
            "month": _integer(value("J8", "/era_time/month")),
            "day": _integer(value("L8", "/era_time/day")),
            "time": _text(value("N8", "/era_time/time")),
        }
        vitals = {
            "hp": {
                "current": _integer(value("E10", "/vitals/hp/current")),
                "max": _integer(value("G10", "/vitals/hp/max")),
                "status": _text(value("I11", "/vitals/hp/status")),
                "major_wound_threshold": _integer(value("D12", "/vitals/hp/major_wound_threshold")),
                "temporary": _integer(value("I12", "/vitals/hp/temporary")),
            },
            "san": {
                "current": _integer(value("N10", "/vitals/san/current")),
                "max": _integer(value("P10", "/vitals/san/max")),
                "status": _text(value("R11", "/vitals/san/status")),
                "daily_loss": _integer(value("N12", "/vitals/san/daily_loss")),
                "remaining": _integer(value("R12", "/vitals/san/remaining")),
            },
            "mp": {
                "current": _integer(value("W10", "/vitals/mp/current")),
                "max": _integer(value("Y10", "/vitals/mp/max")),
                "hourly_recovery": _integer(value("AA11", "/vitals/mp/hourly_recovery")),
            },
            "move": _integer(value("AF10", "/vitals/move")),
            "movement_mode": _text(value("AE12", "/vitals/movement_mode")),
            "damage_bonus": _text(value("AP53", "/vitals/damage_bonus")),
            "build": _integer(value("AP56", "/vitals/build")),
            "dodge": {
                "regular": _integer(value("AP58", "/vitals/dodge/regular")),
                "hard": _integer(value("AR58", "/vitals/dodge/hard")),
                "extreme": _integer(value("AR59", "/vitals/dodge/extreme")),
            },
            "armor": {
                "type": _text(value("AN12", "/vitals/armor/type")),
                "reduction": _text(value("AI12", "/vitals/armor/reduction")),
                "covered_parts": _text(value("AR10", "/vitals/armor/covered_parts")),
            },
        }

        skills: dict[str, dict[str, Any]] = {}
        blocks = (
            ("B", "D", "F", "H", ("J", "L", "N", "P"), "R", "T", "V"),
            ("X", "Z", "AB", "AD", ("AF", "AH", "AJ", "AL"), "AN", "AP", "AR"),
        )
        component_names = ("initial", "growth", "occupation", "interest")
        for row in range(17, 51):
            for (
                mark_col,
                career_col,
                name_col,
                custom_col,
                component_cols,
                reg_col,
                hard_col,
                ext_col,
            ) in blocks:
                base_name = _text(sheet[f"{name_col}{row}"].value)
                custom_name = _text(sheet[f"{custom_col}{row}"].value)
                name = _skill_name(base_name, custom_name)
                if not name:
                    continue
                components = {
                    key: _integer(sheet[f"{column}{row}"].value) or 0
                    for key, column in zip(component_names, component_cols, strict=True)
                }
                regular = _integer(sheet[f"{reg_col}{row}"].value)
                if regular is None:
                    regular = sum(components.values())
                unique_name = name
                suffix = 2
                while unique_name in skills:
                    unique_name = f"{name} ({suffix})"
                    suffix += 1
                pointer = f"/skills/{unique_name}"
                provenance[pointer] = [
                    f"{mark_col}{row}",
                    f"{career_col}{row}",
                    f"{name_col}{row}",
                    f"{custom_col}{row}",
                    *(f"{column}{row}" for column in component_cols),
                    f"{reg_col}{row}",
                    f"{hard_col}{row}",
                    f"{ext_col}{row}",
                ]
                skills[unique_name] = {
                    "name": unique_name,
                    **components,
                    "regular": regular,
                    "hard": _integer(sheet[f"{hard_col}{row}"].value) or regular // 2,
                    "extreme": _integer(sheet[f"{ext_col}{row}"].value) or regular // 5,
                    "improvement_mark": _text(sheet[f"{mark_col}{row}"].value),
                    "occupation_mark": _text(sheet[f"{career_col}{row}"].value),
                }

        weapons: list[dict[str, Any]] = []
        for row in range(54, 58):
            name = _text(sheet[f"B{row}"].value)
            weapon_type = _text(sheet[f"G{row}"].value)
            skill = _text(sheet[f"M{row}"].value)
            if not any((name, weapon_type, skill)) or skill == "←请选择类型":
                continue
            weapon = {
                "id": f"weapon-{row}",
                "name": name,
                "type": weapon_type,
                "skill": skill,
                "regular": _integer(sheet[f"Q{row}"].value),
                "hard": _integer(sheet[f"S{row}"].value),
                "extreme": _integer(sheet[f"U{row}"].value),
                "damage": _text(sheet[f"W{row}"].value),
                "range": _text(sheet[f"AA{row}"].value),
                "impale": _text(sheet[f"AC{row}"].value),
                "attacks": _text(sheet[f"AE{row}"].value),
                "ammo": _text(sheet[f"AG{row}"].value),
                "malfunction": _text(sheet[f"AJ{row}"].value),
            }
            weapons.append(_prune(weapon))
            provenance[f"/weapons/{len(weapons) - 1}"] = [
                f"{column}{row}"
                for column in ("B", "G", "M", "Q", "S", "U", "W", "AA", "AC", "AE", "AG", "AJ")
            ]

        assets = {
            "credit_rating": _text(value("B63", "/assets/credit_rating")),
            "living_standard": _text(value("F63", "/assets/living_standard")),
            "spending_level": _text(value("I63", "/assets/spending_level")),
            "other_assets_summary": _text(value("L63", "/assets/other_assets_summary")),
            "cash": _text(value("O63", "/assets/cash")),
            "currency": _text(value("S63", "/assets/currency")),
            "living_description": _text(value("B64", "/assets/living_description")),
            "financial_description": _text(value("L64", "/assets/financial_description")),
            "categories": [
                _prune(
                    {
                        "type": label,
                        "description": _text(sheet[f"{column}71"].value),
                        "value": _text(sheet[f"{column}76"].value),
                    }
                )
                for label, column in (
                    ("vehicle", "B"),
                    ("residence", "F"),
                    ("luxury", "J"),
                    ("securities", "N"),
                    ("other", "R"),
                )
                if _text(sheet[f"{column}71"].value) or _text(sheet[f"{column}76"].value)
            ],
        }
        background = {
            "appearance": _text(value("AA62", "/background/appearance")),
            "beliefs": _text(value("AA64", "/background/beliefs")),
            "important_people": _text(value("AA66", "/background/important_people")),
            "meaningful_locations": _text(value("AA68", "/background/meaningful_locations")),
            "treasured_possessions": _text(value("AA70", "/background/treasured_possessions")),
            "secrets": _text(value("AA72", "/background/secrets")),
            "traits": _text(value("AA74", "/background/traits")),
            "injuries": _text(value("AA76", "/background/injuries")),
            "phobias_and_manias": _text(value("AA78", "/background/phobias_and_manias")),
            "backstory": _text(value("W80", "/background/backstory")),
        }

        inventory: list[dict[str, Any]] = []
        for row in range(80, 96):
            name = _text(sheet[f"F{row}"].value)
            backpack = _text(sheet[f"N{row}"].value)
            if name:
                inventory.append(
                    _prune(
                        {
                            "id": f"inventory-{row}",
                            "name": name,
                            "location": _text(sheet[f"D{row}"].value),
                            "visibility": _text(sheet[f"B{row}"].value),
                        }
                    )
                )
                provenance[f"/inventory/{len(inventory) - 1}"] = [
                    f"B{row}",
                    f"D{row}",
                    f"F{row}",
                ]
            if backpack:
                inventory.append(
                    {
                        "id": f"backpack-{row}",
                        "name": backpack,
                        "location": "背包",
                        "kind": "contents",
                    }
                )
                provenance[f"/inventory/{len(inventory) - 1}"] = [f"N{row}"]

        experiences: list[dict[str, Any]] = []
        myth_contacts: list[dict[str, Any]] = []
        final_row = min(sheet.max_row or 140, 140)
        for row in range(98, final_row + 1):
            module = _text(sheet[f"B{row}"].value)
            change = _text(sheet[f"J{row}"].value)
            if module and not module.startswith("例"):
                experiences.append(
                    _prune({"id": f"experience-{row}", "module": module, "changes": change})
                )
            encountered = _text(sheet[f"W{row}"].value)
            result = _text(sheet[f"AA{row}"].value)
            note = _text(sheet[f"AK{row}"].value)
            cumulative = _integer(sheet[f"AR{row}"].value)
            if encountered and not encountered.startswith("例"):
                myth_contacts.append(
                    _prune(
                        {
                            "id": f"myth-{row}",
                            "encountered": encountered,
                            "result": result,
                            "note": note,
                            "cumulative": cumulative,
                        }
                    )
                )

        return cast(
            dict[str, Any],
            _prune(
                {
                    "schema_version": 1,
                    "template_id": TEMPLATE_ID,
                    "character_id": _slug(str(identity["name"])),
                    "identity": identity,
                    "era_time": era_time,
                    "attributes": attributes,
                    "vitals": vitals,
                    "skills": skills,
                    "weapons": weapons,
                    "assets": assets,
                    "background": background,
                    "inventory": inventory,
                    "experiences": experiences,
                    "myth_contacts": myth_contacts,
                    "provenance": provenance,
                },
            ),
        )


def roleplay_view(document: dict[str, Any]) -> dict[str, Any]:
    result = copy.deepcopy(document)
    result.pop("provenance", None)
    result.pop("template_id", None)
    return cast(dict[str, Any], _prune(result))


def _collect_overlay(base: Any, current: Any, path: str = "") -> dict[str, Any]:
    if type(base) is not type(current):
        return {path: copy.deepcopy(current)} if path else {}
    if isinstance(base, dict):
        changes: dict[str, Any] = {}
        for key in set(base) | set(current):
            if key in {"provenance", "template_id", "schema_version", "character_id"}:
                continue
            child = f"{path}/{key.replace('~', '~0').replace('/', '~1')}"
            if key not in base:
                changes[child] = copy.deepcopy(current[key])
            elif key not in current:
                changes[child] = None
            else:
                changes.update(_collect_overlay(base[key], current[key], child))
        return changes
    if isinstance(base, list):
        return {path: copy.deepcopy(current)} if base != current else {}
    return {path: copy.deepcopy(current)} if base != current else {}


def preserve_runtime_overlay(
    old_base: dict[str, Any], old_current: dict[str, Any], new_base: dict[str, Any]
) -> tuple[dict[str, Any], list[str]]:
    """把旧基准到旧当前的变化尽力应用到新基准；不兼容路径形成警告。"""
    from qq_mcp_server.store import _pointer_get, _pointer_remove, _pointer_set

    merged = copy.deepcopy(new_base)
    warnings: list[str] = []
    for path, value in _collect_overlay(old_base, old_current).items():
        if not path:
            continue
        exists, _ = _pointer_get(merged, path)
        try:
            if value is None:
                if exists:
                    _pointer_remove(merged, path)
            elif exists:
                _pointer_set(merged, path, value)
            else:
                warnings.append(f"新人物卡中不存在旧动态字段：{path}")
        except ValueError:
            warnings.append(f"无法保留旧动态字段：{path}")
    return merged, warnings


class CharacterCardService:
    def __init__(self, store: MessageStore, storage_dir: Path) -> None:
        self.store = store
        self.storage_dir = storage_dir
        self.staging_dir = storage_dir / "staging"
        self.current_dir = storage_dir / "current"
        self.staging_dir.mkdir(parents=True, exist_ok=True, mode=0o700)
        self.current_dir.mkdir(parents=True, exist_ok=True, mode=0o700)
        self.parser = FixedCharacterCardParser()

    def stage(self, token: str, source: Path) -> tuple[Path, ParsedCard]:
        digest = hashlib.sha256(token.encode()).hexdigest()
        target = self.staging_dir / f"{digest}.xlsx"
        shutil.copyfile(source, target)
        target.chmod(0o600)
        parsed = self.parser.parse(target)
        return target, ParsedCard(parsed.document, parsed.source_sha256, source.name)

    def preview(self, group_key: str, parsed: ParsedCard) -> dict[str, Any]:
        current = self.store.character(group_key)
        old_name = str(current["current"].get("identity", {}).get("name") or "") if current else ""
        same_name = bool(old_name and old_name == parsed.character_name)
        return {
            "character_name": parsed.character_name,
            "player": parsed.document["identity"].get("player"),
            "occupation": parsed.document["identity"].get("occupation"),
            "skill_count": len(parsed.document.get("skills", {})),
            "weapon_count": len(parsed.document.get("weapons", [])),
            "inventory_count": len(parsed.document.get("inventory", [])),
            "previous_character_name": old_name or None,
            "same_character_name": same_name,
            "default_runtime_policy": "preserve" if same_name else "reset",
        }

    def finalize(
        self,
        group_key: str,
        *,
        staged_path: Path,
        source_filename: str,
        runtime_policy: str,
    ) -> dict[str, Any]:
        parsed = self.parser.parse(staged_path)
        previous = self.store.character(group_key)
        policy = runtime_policy
        if policy == "auto":
            old_name = (
                str(previous["current"].get("identity", {}).get("name") or "") if previous else ""
            )
            policy = "preserve" if old_name == parsed.character_name and old_name else "reset"
        if policy not in {"preserve", "reset"}:
            raise ValueError("运行数据策略必须是 auto、preserve 或 reset")
        warnings: list[str] = []
        current_document = copy.deepcopy(parsed.document)
        if policy == "preserve" and previous:
            current_document, warnings = preserve_runtime_overlay(
                previous["base"], previous["current"], parsed.document
            )

        final_path = self.current_dir / f"{group_key}.xlsx"
        temporary = self.current_dir / f".{group_key}.{secrets.token_hex(4)}.xlsx"
        shutil.copyfile(staged_path, temporary)
        temporary.chmod(0o600)
        os.replace(temporary, final_path)
        result = self.store.replace_character(
            group_key,
            source_filename=source_filename,
            source_sha256=parsed.source_sha256,
            source_path=str(final_path),
            base_card=parsed.document,
            current_card=current_document,
            clear_runtime_data=policy == "reset",
        )
        staged_path.unlink(missing_ok=True)
        return {
            "character_name": parsed.character_name,
            "runtime_policy": policy,
            "warnings": warnings,
            "imported_at": result["imported_at"],
        }
